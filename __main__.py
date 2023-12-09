# adding_data.py

from openpyxl import Workbook
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
from datetime import datetime, timezone
import curses

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

class SheetLetters:
    DATE = "A"
    QUI = "B"
    CONTENU = "C"
    REACTION = "D"
    REPONSE = "E"



def creatematch():
    file1 = open('/Users/louis/prog/regex-test-messages/imessage_export/aryane_lippe@icloud.txt', 'r')
    count = 0
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Date"
    sheet["B1"] = "Qui"
    sheet["C1"] = "Contenu"
    sheet["D1"] = "Réaction?"
    sheet["E1"] = "Réponse?"
    lettercounter = 2

    while True:
        count += 1
    # Get next line from file
        line = file1.readline()
        if (line == "This message responded to an earlier message."):
            sheet[SheetLetters.REPONSE] = "VRAI"

        time_and_date = line[:24]
        #time.sleep(1)
        
        try:
            sleepy = datetime.strptime(time_and_date,
                  '%b %d, %Y  %I:%M:%S %p')
            sheet[SheetLetters.DATE + str(lettercounter)] = sleepy.strftime('%a %d %b %Y, %I:%M%p')
            sheet[SheetLetters.REPONSE + str(lettercounter)] = "FAUX"
            line = file1.readline()
            sheet[SheetLetters.QUI + str(lettercounter)] = line
            line = file1.readline()
            sheet[SheetLetters.CONTENU + str(lettercounter)] = line

            lettercounter += 1
        except:
            try: 
                time_and_date = line[4:28]
                sleepy = datetime.strptime(time_and_date, #code failsafe pour quand meme avoir les données
                  '%b %d, %Y  %I:%M:%S %p')
                sheet[SheetLetters.DATE + str(lettercounter)] = sleepy.strftime('%a %d %b %Y, %I:%M%p')
                sheet[SheetLetters.REPONSE + str(lettercounter)] = "VRAI"
                line = file1.readline()
                name = line[4:] 
                sheet[SheetLetters.QUI + str(lettercounter)] = name
                line = file1.readline()
                name = line[4:] 
                sheet[SheetLetters.CONTENU + str(lettercounter)] = name
                lettercounter += 1
            
            except: 
                if not line:
                    break
                else: 
                    continue
            
        print(bcolors.OKCYAN + str(sleepy) + bcolors.ENDC)

    # if line is empty
    # end of file is reached
        if not line:
            break
        #print(line)
    workbook.save("hello.xlsx")
    file1.close()
    #print(count)


def create_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Date"
    sheet["B1"] = "Qui"
    sheet["C1"] = "Contenu"
    sheet["D1"] = "Réaction?"
    sheet["E1"] = "Réponse?"

    workbook.save(path)


if __name__ == "__main__":
    create_workbook("hello.xlsx")

creatematch()